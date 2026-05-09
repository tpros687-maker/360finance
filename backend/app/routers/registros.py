import base64
import io
import json
import math
import os
import uuid
from datetime import date
from decimal import Decimal
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, select, or_
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.deps import get_current_user
from app.models.categoria import Categoria, TipoMovimiento
from app.models.mapa import Potrero
from app.models.registro import Registro
from app.models.user import User
from app.schemas.registro import (
    ExtraerComprobanteResponse,
    PaginatedRegistros,
    RegistroCreate,
    RegistroRead,
    RegistroUpdate,
    ResumenCategoria,
    ResumenMes,
    ResumenResponse,
)
from app.services.asistente import _get_client as _groq_client
from app.services.rentabilidad import invalidar_cache_potrero

router = APIRouter(prefix="/registros", tags=["registros"])

UPLOADS_DIR = "/app/uploads/comprobantes"
ALLOWED_EXTENSIONS = {"jpg", "jpeg", "png", "pdf"}


def _build_filters(
    query,
    user_id: int,
    tipo: Optional[TipoMovimiento],
    categoria_id: Optional[int],
    potrero_id: Optional[int],
    fecha_desde: Optional[date],
    fecha_hasta: Optional[date],
    q: Optional[str],
):
    query = query.where(Registro.user_id == user_id)
    if tipo:
        query = query.where(Registro.tipo == tipo)
    if categoria_id:
        query = query.where(Registro.categoria_id == categoria_id)
    if potrero_id:
        query = query.where(Registro.potrero_id == potrero_id)
    if fecha_desde:
        query = query.where(Registro.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.where(Registro.fecha <= fecha_hasta)
    if q:
        query = query.where(Registro.descripcion.ilike(f"%{q}%"))
    return query


@router.get("/resumen", response_model=ResumenResponse)
async def get_resumen(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> ResumenResponse:
    """Totales globales, por categoría y por mes. Para el Dashboard."""
    uid = current_user.id

    gastos_q = await db.execute(
        select(func.coalesce(func.sum(Registro.monto), 0)).where(
            Registro.user_id == uid, Registro.tipo == TipoMovimiento.gasto
        )
    )
    ingresos_q = await db.execute(
        select(func.coalesce(func.sum(Registro.monto), 0)).where(
            Registro.user_id == uid, Registro.tipo == TipoMovimiento.ingreso
        )
    )
    total_gastos = Decimal(str(gastos_q.scalar()))
    total_ingresos = Decimal(str(ingresos_q.scalar()))

    cat_q = await db.execute(
        select(
            Registro.categoria_id,
            Categoria.nombre,
            Categoria.tipo,
            Categoria.color,
            func.sum(Registro.monto).label("total"),
        )
        .join(Categoria, Registro.categoria_id == Categoria.id)
        .where(Registro.user_id == uid)
        .group_by(Registro.categoria_id, Categoria.nombre, Categoria.tipo, Categoria.color)
        .order_by(func.sum(Registro.monto).desc())
    )
    por_categoria = [
        ResumenCategoria(
            categoria_id=row.categoria_id,
            nombre=row.nombre,
            tipo=row.tipo,
            color=row.color,
            total=Decimal(str(row.total)),
        )
        for row in cat_q.all()
    ]

    mes_q = await db.execute(
        select(
            func.to_char(Registro.fecha, "YYYY-MM").label("mes"),
            func.sum(
                case((Registro.tipo == TipoMovimiento.gasto, Registro.monto), else_=0)
            ).label("gastos"),
            func.sum(
                case((Registro.tipo == TipoMovimiento.ingreso, Registro.monto), else_=0)
            ).label("ingresos"),
        )
        .where(Registro.user_id == uid)
        .group_by("mes")
        .order_by("mes")
    )
    por_mes = [
        ResumenMes(
            mes=row.mes,
            gastos=Decimal(str(row.gastos)),
            ingresos=Decimal(str(row.ingresos)),
        )
        for row in mes_q.all()
    ]

    return ResumenResponse(
        total_gastos=total_gastos,
        total_ingresos=total_ingresos,
        balance=total_ingresos - total_gastos,
        por_categoria=por_categoria,
        por_mes=por_mes,
    )


@router.get("/exportar")
async def exportar_registros(
    formato: str = Query(..., description="excel o pdf"),
    tipo: Optional[TipoMovimiento] = Query(None),
    categoria_id: Optional[int] = Query(None),
    potrero_id: Optional[int] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    q: Optional[str] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Exporta los registros filtrados en Excel o PDF."""
    uid = current_user.id

    items_q = _build_filters(
        select(Registro), uid, tipo, categoria_id, potrero_id, fecha_desde, fecha_hasta, q
    ).order_by(Registro.fecha.desc(), Registro.id.desc())

    items = list((await db.execute(items_q)).scalars().all())

    if formato == "excel":
        return _export_excel(items)
    elif formato == "pdf":
        return _export_pdf(items)
    else:
        raise HTTPException(status_code=400, detail="Formato debe ser 'excel' o 'pdf'")


def _export_excel(items: list) -> StreamingResponse:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registros"

    # Header
    headers = ["Fecha", "Tipo", "Categoría", "Potrero", "Descripción", "Monto"]
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 16

    # Data
    for row_idx, r in enumerate(items, 2):
        ws.cell(row=row_idx, column=1, value=str(r.fecha))
        ws.cell(row=row_idx, column=2, value=r.tipo.value if hasattr(r.tipo, "value") else str(r.tipo))
        ws.cell(row=row_idx, column=3, value=r.categoria.nombre if r.categoria else "")
        ws.cell(row=row_idx, column=4, value=r.potrero.nombre if r.potrero else "")
        ws.cell(row=row_idx, column=5, value=r.descripcion or "")
        monto_cell = ws.cell(row=row_idx, column=6, value=float(r.monto))
        monto_cell.number_format = '#,##0.00'
        if str(r.tipo) == "gasto":
            monto_cell.font = Font(color="EF4444")
        else:
            monto_cell.font = Font(color="10B981")

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=registros.xlsx"},
    )


def _export_pdf(items: list) -> StreamingResponse:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=15 * mm,
        leftMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
    )

    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph("Registros financieros", styles["Title"]))
    elements.append(Spacer(1, 6 * mm))

    headers = ["Fecha", "Tipo", "Categoría", "Potrero", "Descripción", "Monto"]
    data = [headers]

    for r in items:
        tipo_str = r.tipo.value if hasattr(r.tipo, "value") else str(r.tipo)
        data.append([
            str(r.fecha),
            tipo_str.capitalize(),
            r.categoria.nombre if r.categoria else "",
            r.potrero.nombre if r.potrero else "",
            (r.descripcion or "")[:60],
            f"${float(r.monto):,.2f}",
        ])

    col_widths = [28 * mm, 22 * mm, 45 * mm, 40 * mm, 90 * mm, 32 * mm]
    table = Table(data, colWidths=col_widths, repeatRows=1)

    style = TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E293B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F8FAFC")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ])

    # Color monto por tipo
    for i, r in enumerate(items, 1):
        tipo_str = r.tipo.value if hasattr(r.tipo, "value") else str(r.tipo)
        color = colors.HexColor("#EF4444") if tipo_str == "gasto" else colors.HexColor("#10B981")
        style.add("TEXTCOLOR", (5, i), (5, i), color)

    table.setStyle(style)
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=registros.pdf"},
    )


@router.get("", response_model=PaginatedRegistros)
async def list_registros(
    tipo: Optional[TipoMovimiento] = Query(None),
    categoria_id: Optional[int] = Query(None),
    potrero_id: Optional[int] = Query(None),
    fecha_desde: Optional[date] = Query(None),
    fecha_hasta: Optional[date] = Query(None),
    q: Optional[str] = Query(None, description="Buscar por descripción"),
    page: int = Query(1, ge=1),
    limit: int = Query(20, ge=1, le=100),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> PaginatedRegistros:
    uid = current_user.id

    count_q = _build_filters(
        select(func.count()).select_from(Registro), uid, tipo, categoria_id, potrero_id, fecha_desde, fecha_hasta, q
    )
    total = (await db.execute(count_q)).scalar() or 0

    items_q = _build_filters(
        select(Registro), uid, tipo, categoria_id, potrero_id, fecha_desde, fecha_hasta, q
    ).order_by(Registro.fecha.desc(), Registro.id.desc()).offset((page - 1) * limit).limit(limit)

    items = list((await db.execute(items_q)).scalars().all())

    return PaginatedRegistros(
        items=items,
        total=total,
        page=page,
        limit=limit,
        pages=max(1, math.ceil(total / limit)),
    )


@router.post("", response_model=RegistroRead, status_code=status.HTTP_201_CREATED)
async def create_registro(
    payload: RegistroCreate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Registro:
    await _assert_categoria_accessible(payload.categoria_id, current_user.id, db)
    if payload.potrero_id is not None:
        await _assert_potrero_accessible(payload.potrero_id, current_user.id, db)

    registro = Registro(
        user_id=current_user.id,
        categoria_id=payload.categoria_id,
        potrero_id=payload.potrero_id,
        tipo=payload.tipo,
        monto=payload.monto,
        moneda=payload.moneda,
        fecha=payload.fecha,
        descripcion=payload.descripcion,
        comprobante_url=payload.comprobante_url,
        tipo_imputacion=payload.tipo_imputacion,
        actividad_tipo=payload.actividad_tipo,
        actividad_id=payload.actividad_id,
    )
    db.add(registro)
    await db.commit()
    if payload.potrero_id is not None:
        await invalidar_cache_potrero(payload.potrero_id, db)
    await db.refresh(registro)
    result = await db.execute(select(Registro).where(Registro.id == registro.id))
    return result.scalar_one()


_MIME_MAP: dict[str, str] = {
    "jpg": "image/jpeg",
    "jpeg": "image/jpeg",
    "png": "image/png",
    "pdf": "application/pdf",
}

_EXTRACCION_PROMPT = (
    'Analizá esta factura o comprobante y extraé la información en JSON válido sin texto extra:\n'
    '{"monto": número o null, "proveedor": string o null, "fecha": "YYYY-MM-DD" o null, '
    '"descripcion": string breve o null, "categoria_sugerida": string o null}'
)


def _confianza(data: dict) -> str:
    filled = sum(1 for k in ("monto", "proveedor", "fecha") if data.get(k) is not None)
    if filled >= 3:
        return "alta"
    if filled >= 2:
        return "media"
    return "baja"


@router.post("/extraer-comprobante", response_model=ExtraerComprobanteResponse)
async def extraer_comprobante(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
) -> ExtraerComprobanteResponse:
    """Extrae datos de una factura o comprobante usando visión IA."""
    _empty = ExtraerComprobanteResponse(
        monto=None, proveedor=None, fecha=None,
        descripcion=None, categoria_sugerida=None, confianza="baja",
    )

    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Formato no permitido. Usá: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    try:
        content = await file.read()
        b64 = base64.b64encode(content).decode("utf-8")
        mime = _MIME_MAP.get(ext, "image/jpeg")

        client = _groq_client()
        response = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {
                            "type": "image_url",
                            "image_url": {"url": f"data:{mime};base64,{b64}"},
                        },
                        {"type": "text", "text": _EXTRACCION_PROMPT},
                    ],
                }
            ],
            max_tokens=512,
            temperature=0.1,
        )

        raw = (response.choices[0].message.content or "").strip()
        if raw.startswith("```"):
            parts = raw.split("```", 2)
            raw = parts[1]
            if raw.startswith("json"):
                raw = raw[4:]
            if len(parts) > 2:
                raw = raw.rsplit("```", 1)[0]
        raw = raw.strip()

        data: dict = json.loads(raw)
        return ExtraerComprobanteResponse(
            monto=float(data["monto"]) if data.get("monto") is not None else None,
            proveedor=data.get("proveedor"),
            fecha=data.get("fecha"),
            descripcion=data.get("descripcion"),
            categoria_sugerida=data.get("categoria_sugerida"),
            confianza=_confianza(data),
        )
    except Exception:
        return _empty


@router.put("/{registro_id}", response_model=RegistroRead)
async def update_registro(
    registro_id: int,
    payload: RegistroUpdate,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Registro:
    registro = await _get_own_registro(registro_id, current_user.id, db)
    old_potrero_id = registro.potrero_id

    if payload.categoria_id is not None:
        await _assert_categoria_accessible(payload.categoria_id, current_user.id, db)
        registro.categoria_id = payload.categoria_id
    if payload.tipo is not None:
        registro.tipo = payload.tipo
    if payload.monto is not None:
        registro.monto = payload.monto
    if payload.moneda is not None:
        registro.moneda = payload.moneda
    if payload.fecha is not None:
        registro.fecha = payload.fecha
    if payload.descripcion is not None:
        registro.descripcion = payload.descripcion
    if payload.comprobante_url is not None:
        registro.comprobante_url = payload.comprobante_url
    # potrero_id puede setearse a None explícitamente para desasociar
    if "potrero_id" in payload.model_fields_set:
        if payload.potrero_id is not None:
            await _assert_potrero_accessible(payload.potrero_id, current_user.id, db)
        registro.potrero_id = payload.potrero_id
    if "tipo_imputacion" in payload.model_fields_set:
        registro.tipo_imputacion = payload.tipo_imputacion
    if "actividad_tipo" in payload.model_fields_set:
        registro.actividad_tipo = payload.actividad_tipo
    if "actividad_id" in payload.model_fields_set:
        registro.actividad_id = payload.actividad_id

    await db.commit()
    for pid in {p for p in [old_potrero_id, registro.potrero_id] if p is not None}:
        await invalidar_cache_potrero(pid, db)
    result = await db.execute(select(Registro).where(Registro.id == registro.id))
    return result.scalar_one()


@router.post("/{registro_id}/comprobante", response_model=RegistroRead)
async def upload_comprobante(
    registro_id: int,
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> Registro:
    registro = await _get_own_registro(registro_id, current_user.id, db)

    ext = (file.filename or "").rsplit(".", 1)[-1].lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Formato no permitido. Usá: {', '.join(ALLOWED_EXTENSIONS)}",
        )

    os.makedirs(UPLOADS_DIR, exist_ok=True)
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = os.path.join(UPLOADS_DIR, filename)

    content = await file.read()
    with open(filepath, "wb") as f:
        f.write(content)

    registro.comprobante_url = f"/uploads/comprobantes/{filename}"
    await db.commit()
    result = await db.execute(select(Registro).where(Registro.id == registro.id))
    return result.scalar_one()


@router.delete("/{registro_id}", status_code=status.HTTP_204_NO_CONTENT)
async def delete_registro(
    registro_id: int,
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
) -> None:
    registro = await _get_own_registro(registro_id, current_user.id, db)
    potrero_id = registro.potrero_id
    await db.delete(registro)
    await db.commit()
    if potrero_id is not None:
        await invalidar_cache_potrero(potrero_id, db)


# ── Helpers ──────────────────────────────────────────────────────────────────

async def _get_own_registro(registro_id: int, user_id: int, db: AsyncSession) -> Registro:
    result = await db.execute(select(Registro).where(Registro.id == registro_id))
    registro = result.scalar_one_or_none()
    if registro is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Registro no encontrado")
    if registro.user_id != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Sin permisos sobre este registro")
    return registro


async def _assert_categoria_accessible(categoria_id: int, user_id: int, db: AsyncSession) -> None:
    result = await db.execute(select(Categoria).where(Categoria.id == categoria_id))
    cat = result.scalar_one_or_none()
    if cat is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Categoría no encontrada")
    if cat.es_personalizada and cat.user_id != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="No podés usar esa categoría")


async def _assert_potrero_accessible(potrero_id: int, user_id: int, db: AsyncSession) -> None:
    result = await db.execute(select(Potrero).where(Potrero.id == potrero_id))
    potrero = result.scalar_one_or_none()
    if potrero is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Potrero no encontrado")
    if potrero.user_id != user_id:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Sin permisos sobre ese potrero")
